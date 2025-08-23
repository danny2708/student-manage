# app/services/sheet_import_user_service.py
from fastapi import UploadFile
from sqlalchemy.orm import Session
from openpyxl import load_workbook # type: ignore
from io import BytesIO
from app.crud import excel_import_user_crud


def import_users(file: UploadFile, db: Session):
    try:
        contents = file.file.read()  # đọc sync
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)

        # --- Đọc sheet Student ---
        if "Student" not in workbook.sheetnames:
            raise ValueError("Excel file must contain a 'Student' sheet")
        ws_student = workbook["Student"]
        student_rows = []
        for row in ws_student.iter_rows(min_row=2, values_only=True):
            row = row[1:]  # bỏ cột A (STT)
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            if clean_row[0] and "@" in clean_row[0]:
                student_rows.append(clean_row)

        email_to_student_id = excel_import_user_crud.import_students(db, student_rows)

        # --- Đọc sheet Parent ---
        if "Parent" not in workbook.sheetnames:
            raise ValueError("Excel file must contain a 'Parent' sheet")
        ws_parent = workbook["Parent"]
        parent_rows = []
        for row in ws_parent.iter_rows(min_row=2, values_only=True):
            row = row[1:]
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            if clean_row[0] and "@" in clean_row[0]:
                parent_rows.append(clean_row)

        email_to_parent_id = excel_import_user_crud.import_parents(
            db, parent_rows, email_to_student_id
        )

        return {
            "students": email_to_student_id,
            "parents": email_to_parent_id,
        }

    except Exception as e:
        raise RuntimeError(f"Import failed: {str(e)}")
