# app/services/excel_test_service.py
from fastapi import UploadFile
from sqlalchemy.orm import Session
from openpyxl import load_workbook  # type: ignore
from io import BytesIO
from app.models.class_model import Class
from app.schemas.test_schema import TestCreate
from app.crud import test_crud
from .. import service_helper

def import_tests_from_excel(db: Session, file: UploadFile, class_id: int):
    try:
        contents = file.file.read()
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)

        ws = workbook.active  # sheet đầu tiên
        test_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row: [STT, test_name, student_id, student_name, class, subject, score, exam_date]
            test_name = (row[1] or "").strip() if row[1] else ""
            student_id = int(row[2]) if row[2] else None
            score = float(row[6]) if row[6] is not None else None
            exam_date = service_helper.parse_date_safe(row[7])

            if not (test_name and student_id and score is not None and exam_date):
                continue

            test_rows.append({
                "test_name": test_name,
                "student_id": student_id,
                "score": score,
                "exam_date": exam_date,
            })

        # Lấy thông tin class để gán teacher_id & subject_id
        db_class = db.query(Class).filter(Class.class_id == class_id).first()
        if not db_class:
            raise ValueError(f"Class with id {class_id} not found")

        created_tests = []
        for r in test_rows:
            test_in = TestCreate(
                test_name=r["test_name"],
                student_id=r["student_id"],
                class_id=class_id,
                subject_id=db_class.subject_id,
                teacher_id=db_class.teacher_id,
                score=r["score"],
                exam_date=r["exam_date"],
            )
            db_test = test_crud.create_test(db, test_in)
            created_tests.append(db_test)

        return {"imported": len(created_tests)}

    except Exception as e:
        raise RuntimeError(f"Import tests failed: {str(e)}")
