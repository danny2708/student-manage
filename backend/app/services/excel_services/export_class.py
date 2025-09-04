from sqlalchemy.orm import Session
from openpyxl import Workbook  # type: ignore
from fastapi.responses import StreamingResponse
from io import BytesIO

from app.models.class_model import Class
from app.models.teacher_model import Teacher
from app.models.student_model import Student
from app.models.user_model import User
from app.models.enrollment_model import Enrollment, EnrollmentStatus

def export_class(db: Session, class_id: int):
    """
    Xuất dữ liệu lớp học ra Excel:
    - B1: class_name
    - D1: teacher full_name
    - B3-D3: header
    - Dữ liệu student lấy từ bảng Enrollment
    """

    # --- Truy vấn thông tin class ---
    db_class = db.query(Class).filter(Class.class_id == class_id).first()
    if not db_class:
        raise ValueError(f"Class id={class_id} not found")

    # --- Truy vấn teacher ---
    teacher_name = ""
    if db_class.teacher_user_id:
        teacher_user = (
            db.query(User)
            .join(Teacher, Teacher.user_id == User.user_id)
            .filter(Teacher.user_id == db_class.teacher_user_id)
            .first()
        )
        teacher_name = teacher_user.full_name if teacher_user else ""

    # --- Truy vấn students qua Enrollment ---
    students = (
        db.query(Student.user_id, User.full_name, User.date_of_birth)
        .join(User, User.user_id == Student.user_id)
        .join(Enrollment, Enrollment.student_user_id == Student.user_id)
        .filter(
            Enrollment.class_id == class_id,
            Enrollment.enrollment_status == EnrollmentStatus.active  # chỉ lấy học sinh đang active
        )
        .all()
    )

    # --- Tạo file Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header thông tin class
    ws["A1"] = "Class:"
    ws["B1"] = db_class.class_name
    ws["C1"] = "Teacher:"
    ws["D1"] = teacher_name

    # Header bảng students
    ws["A3"] = "STT"
    ws["B3"] = "student_user_id"
    ws["C3"] = "date_of_birth"
    ws["D3"] = "student_name"

    # Ghi dữ liệu students
    for idx, st in enumerate(students, start=1):
        ws.cell(row=3 + idx, column=1, value=idx)                   # STT
        ws.cell(row=3 + idx, column=2, value=st.student_user_id)    # student_user_id
        ws.cell(row=3 + idx, column=3, value=st.date_of_birth)      # dob
        ws.cell(row=3 + idx, column=4, value=st.full_name)          # name

    # Xuất ra response
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"class_{class_id}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
